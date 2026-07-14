# -*- coding: utf-8 -*-
"""
폐기물 입찰 레이더 ── v2.7
변경이력:
  v2.0  5개 기관 통합
  v2.6  국방부 원본 로직 반영 (service_key 언더스코어, 날짜필터 없음)
  v2.7  흰화면 오류 수정 (구글폰트/복잡CSS 제거), 누락 함수 복구
"""
import streamlit as st
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception:
    pass
import requests
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from urllib.parse import unquote
import re, io
from concurrent.futures import ThreadPoolExecutor, as_completed

st.set_page_config(
    page_title="SWEEP · 입찰공고 통합 수집",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── 최소 CSS (구글폰트 제거 → 흰화면 방지) ──
st.markdown("""
<style>
[data-testid="stSidebar"] { background-color: #1e2235; }
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div { color: #dde3f0 !important; }
.stButton > button {
    background: #2563eb !important; color: white !important;
    border: none !important; font-weight: 600 !important;
}
.stButton > button:hover { background: #1d4ed8 !important; }
[data-testid="stDownloadButton"] button {
    border: 1px solid #2563eb !important; color: #2563eb !important;
    background: white !important;
}
</style>
""", unsafe_allow_html=True)

# =====================================================================
# 0. 공통 설정
# =====================================================================
SERVICE_KEY = '9ada16f8e5bc00e68aa27ceaa5a0c2ae3d4a5e0ceefd9fdca653b03da27eebf0'
HEADERS     = {'User-Agent': 'Mozilla/5.0'}
VERSION     = "v4.5"
TODAY       = datetime.now()
SCSBID_URL  = 'http://apis.data.go.kr/1230000/as/ScsbidInfoService/getOpengResultListInfoServcPPSSrch'

DEFAULT_KEYWORDS = ["폐기물", "운반", "폐목재", "폐합성수지", "잔재", "가연성", "낙엽",
                    "식물성", "부유", "초본류", "초목류", "임목", "폐가구",
                    "대형", "생활", "수거", "우드칩", "운수", "재활용", "톱밥", "적환장", "방치"]
# 기본값 - 사이드바에서 사용자가 편집 가능
DEFAULT_LICENSES     = ['1226', '1227', '6786', '6770']
DEFAULT_LH_LICENSES  = ['폐기물처리업', '폐기물수집운반업', '폐기물재활용업']
# 런타임 값 (사이드바에서 덮어씀)
OUR_LICENSES        = DEFAULT_LICENSES
WASTE_LICENSE_NAMES = DEFAULT_LH_LICENSES
KWATER_DETAIL_BASE  = "https://ebid.kwater.or.kr/wq/index.do?w2xPath=/ui/index.xml&view=bidpblanc/bidpblancsttus/BIDBD32000002.xml&tndrPbanno="
KOGAS_HOME = "https://k-ebid.kogas.or.kr"
D2B_HOME   = "https://www.d2b.go.kr/"
LH_HOME    = "https://ebid.lh.or.kr/"

# =====================================================================
# 1. 공통 함수
# =====================================================================
def _region_token_pass(token):
    token = token.strip()
    if not token: return True
    if any(f in token for f in ['전국', '제한없음']): return True
    if '평택' in token or '화성' in token: return True
    if token == '경기도': return True
    return False

def region_pass(region_text, test_mode=False):
    if test_mode: return True
    if region_text is None: return True
    tokens = list(region_text) if isinstance(region_text, (list, tuple, set)) \
             else re.split(r'[,]', str(region_text).strip())
    return any(_region_token_pass(t) for t in tokens) if tokens else True

def license_code_pass(license_text, test_mode=False):
    if test_mode: return True
    if not license_text: return True
    return any(code in str(license_text) for code in OUR_LICENSES)

def lh_clean(text):
    return re.sub(r'<!\[CDATA\[|\]\]>', '', text).strip() if text else ""

def lh_license_pass(item, test_mode=False):
    if test_mode: return True
    for n in range(1, 11):
        slots = [lh_clean(item.findtext(f'req{n}Reqlic{m}Nm')) for m in range(1, 11)]
        slots = [s for s in slots if s]
        if not slots: continue
        if not any(name in s for s in slots for name in WASTE_LICENSE_NAMES):
            return False
    return True

def format_d2b_dt(val):
    if not val or str(val).strip() in ('', '-'): return '-'
    s = str(val).replace('.0', '').strip()
    try:
        if len(s) >= 12: return f"{s[:4]}-{s[4:6]}-{s[6:8]} {s[8:10]}:{s[10:12]}"
        elif len(s) >= 8: return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
        return s
    except: return s

def to_row(source, notice_no, title, agency, notice_dt, close_dt, open_dt,
           amount, region, license_info, keyword, url='', extra='', lwlt_rate='-'):
    return {
        '출처기관': source, '공고번호': notice_no, '공고명': title,
        '수요/발주기관': agency, '마감일시': close_dt, '금액(원)': amount,
        '낙찰하한율': lwlt_rate, '지역제한': region,
        '면허정보': license_info, '상세URL': url,
        # 내부용 (표시 안 함, 엑셀 다운로드용)
        '_공고일시': notice_dt, '_개찰일시': open_dt, '_키워드': keyword, '_비고': extra,
    }




def fetch_scsbid_rates_safe(keywords_tuple: tuple) -> dict:
    """
    A방식: 역사적 공고별 투찰율/낙찰하한율 비율 평균
    1) 낙찰정보 API로 과거 3개월 개찰결과 + 키워드 매칭
    2) 매칭 공고의 낙찰하한율을 입찰공고 API로 추가 조회
    3) ratio = 투찰율/낙찰하한율 평균 = 참고값
    """
    import xml.etree.ElementTree as ET
    from collections import defaultdict
    from concurrent.futures import ThreadPoolExecutor, as_completed as asc
    try:
        kw_sorted = sorted(keywords_tuple, key=len, reverse=True)
        BID_URL = 'https://apis.data.go.kr/1230000/ad/BidPublicInfoService/getBidPblancListInfoServcPPSSrch'
        KEY = unquote(SERVICE_KEY)

        # Step 1: 3개월 개찰결과 수집
        all_items = []
        for m in range(3):
            e_dt = TODAY - timedelta(days=30 * m)
            s_dt = TODAY - timedelta(days=30 * (m + 1))
            try:
                res = requests.get(SCSBID_URL, params={
                    'ServiceKey': KEY, 'inqryDiv': '1',
                    'inqryBgnDt': s_dt.strftime('%Y%m%d') + '0000',
                    'inqryEndDt': e_dt.strftime('%Y%m%d') + '2359',
                    'numOfRows': '300', 'pageNo': '1',
                }, timeout=10)
                if res.status_code != 200: continue
                root = ET.fromstring(res.text)
                if root.findtext('.//resultCode') != '00': continue
                all_items.extend(root.findall('.//item'))
            except Exception: continue

        # Step 2: 키워드 매칭 + (bidNtceNo, 투찰율) 추출
        matched_bids = []
        for item in all_items:
            bid_nm = item.findtext('bidNtceNm') or ''
            oci    = item.findtext('opengCorpInfo') or ''
            if '완료' not in (item.findtext('progrsDivCdNm') or ''): continue
            kw = next((k for k in kw_sorted if k in bid_nm), None)
            if not kw: continue
            bid_no = item.findtext('bidNtceNo') or ''
            if not bid_no: continue
            parts = oci.split('^')
            if len(parts) >= 5:
                try:
                    t_rate = float(parts[4])
                    if 50 < t_rate < 100:
                        matched_bids.append((bid_no, t_rate, kw))
                except Exception: pass

        if not matched_bids:
            return {}

        # Step 3: 낙찰하한율 일괄 조회
        # 같은 날짜 범위로 나라장터 API 조회 → bidNtceNo 매칭
        target = matched_bids[:40]
        target_nos = {b[0] for b in target}
        lwlt_map = {}

        # 각 월별 날짜 범위로 나라장터 입찰공고 조회
        for m in range(3):
            e_dt = TODAY - timedelta(days=30 * m)
            s_dt = TODAY - timedelta(days=30 * (m + 1))
            try:
                res = requests.get(BID_URL, params={
                    'serviceKey': KEY, 'numOfRows': '500', 'pageNo': '1',
                    'type': 'json', 'inqryDiv': '1',
                    'inqryBgnDt': s_dt.strftime('%Y%m%d'),
                    'inqryEndDt': e_dt.strftime('%Y%m%d'),
                }, timeout=10)
                if res.status_code != 200: continue
                items = res.json().get('response',{}).get('body',{}).get('items',[])
                if isinstance(items, dict): items = [items]
                for it in items:
                    bno = str(it.get('bidNtceNo',''))
                    if bno in target_nos:
                        lwlt = it.get('sucsfbidLwltRate')
                        if lwlt:
                            try: lwlt_map[bno] = float(lwlt)
                            except Exception: pass
                if len(lwlt_map) >= len(target_nos): break
            except Exception: continue

        # Step 4: ratio = 투찰율 / 낙찰하한율
        kw_ratios = defaultdict(list)
        for bid_no, t_rate, kw in target:
            lwlt = lwlt_map.get(bid_no)
            if lwlt and lwlt > 0:
                ratio = (t_rate / 100) / (lwlt / 100)
                if 0.85 < ratio < 1.15:
                    kw_ratios[kw].append({'ratio': ratio, 't_rate': t_rate, 'lwlt': lwlt})

        result = {}
        all_r = [d['ratio'] for v in kw_ratios.values() for d in v]
        for kw, vals in kw_ratios.items():
            ratios = [d['ratio'] for d in vals]
            result[kw] = {
                'ratio':      round(sum(ratios)/len(ratios), 4),
                'count':      len(ratios),
                'avg_t_rate': round(sum(d['t_rate'] for d in vals)/len(vals), 3),
                'avg_lwlt':   round(sum(d['lwlt']   for d in vals)/len(vals), 3),
            }
        if all_r:
            result['__전체__'] = {
                'ratio': round(sum(all_r)/len(all_r), 4),
                'count': len(all_r),
            }
        return result
    except Exception:
        return {}


def make_dajang_excel(sel_df: "pd.DataFrame") -> bytes:
    """
    입찰관리대장 양식 생성 - v3.3
    수정:
    - 눈금선 제거
    - L열 숨기기(기타비용 제거), 수익공식에서 L 제거
    - P,Q,R 셀 2행 merge (비고/수익/수익률)
    - B열 열고정 제거 (행만 고정)
    - No 블록별 굵은 외곽 테두리
    - URL: bidNtceNo 기반 직접 조립
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "입찰관리대장"

    # ── 눈금선 제거 ──
    ws.sheet_view.showGridLines = False

    YELLOW  = PatternFill("solid", fgColor="FFFF00")
    GREEN_H = PatternFill("solid", fgColor="70AD47")
    GREEN_L = PatternFill("solid", fgColor="E2F0D9")
    GRAY    = PatternFill("solid", fgColor="D9D9D9")
    RED_F   = PatternFill("solid", fgColor="FF0000")
    WHITE   = PatternFill("solid", fgColor="FFFFFF")

    thin  = Side(style="thin")
    med   = Side(style="medium")
    thick = Side(style="medium")

    def bd_thin(): return Border(left=thin, right=thin, top=thin, bottom=thin)
    def bd_thick(l=True, r=True, t=True, b=True):
        return Border(
            left   = med if l else thin,
            right  = med if r else thin,
            top    = med if t else thin,
            bottom = med if b else thin,
        )

    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    rgt = Alignment(horizontal="right",  vertical="center")

    # ── 열 너비 (L열 숨기기) ──
    for col, w in {
        "B":4, "C":20, "D":15, "E":14, "F":16,
        "G":14, "H":10, "I":16, "J":11, "K":10,
        "M":13, "N":11, "O":11, "P":18, "Q":13, "R":10
    }.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["L"].hidden = True   # L열 숨기기

    # ── 타이틀 (1행) ──
    ws.merge_cells("B1:R1")
    c = ws["B1"]
    c.value = "입  찰  관  리  대  장"
    c.font  = Font(bold=True, size=16); c.alignment = ctr
    ws.row_dimensions[1].height = 35
    for i in range(2, 7): ws.row_dimensions[i].height = 4

    # ── 헤더 (7행) ──
    HDR = 7
    hdr_items = [
        ("B","No"),("C","공고명"),("D","공고번호"),("E","발주처"),
        ("F","마감일시"),("G","기초금액\n(원)"),("H","폐기물\n성상"),
        ("I","현장주소"),("J","단가"),("K","물량\n(톤)"),
        ("M","계약금액\n(예상)"),("N","낙찰률\n단가"),
        ("O","운반거리\n(편도)"),("P","비고"),("Q","수익"),("R","수익률\n(%)"),
    ]
    for col, val in hdr_items:
        c = ws[f"{col}{HDR}"]
        c.value = val; c.font = Font(bold=True, color="FFFFFF")
        c.fill = GREEN_H; c.alignment = ctr; c.border = bd_thin()
    ws.row_dimensions[HDR].height = 30

    # ── Freeze: 행만 고정 (B열 열고정 제거) ──
    ws.freeze_panes = f"A{HDR+1}"

    # ── 데이터 블록 (3행) ──
    SR    = HDR + 1
    BLOCK = 3
    COL_RANGE = range(2, 19)   # B(2)~R(18)

    for idx, (_, row) in enumerate(sel_df.iterrows()):
        r = SR + idx * BLOCK

        try:
            amt = int(float(str(row.get("금액(원)","0")).replace(",","").replace("원","").strip() or 0))
        except: amt = 0

        lwlt = row.get("낙찰하한율", "-")
        try:
            lwlt_f = float(lwlt) if str(lwlt) not in ("-","","None") else None
        except: lwlt_f = None

        clos   = str(row.get("마감일시","-"))
        bid_no = str(row.get("공고번호",""))
        # URL: fetch 단계에서 이미 올바른 패턴으로 조립됨
        api_url = row.get("상세URL","")
        direct_url = api_url if api_url and api_url not in ("-","") else ""

        ws.row_dimensions[r  ].height = 48
        ws.row_dimensions[r+1].height = 22
        ws.row_dimensions[r+2].height = 25

        # No 셀 (merge는 나중에 일괄 처리)
        c = ws[f"B{r}"]
        c.value = idx+1; c.font = Font(bold=True, size=12); c.alignment = ctr
        c.fill  = WHITE

        # ── r+0: 메인 행 ──────────────────────────────────────────
        def sm(col, val=None, fill=None, fmt=None, al=None):
            c = ws[f"{col}{r}"]
            if val  is not None: c.value = val
            if fill is not None: c.fill  = fill
            if fmt  is not None: c.number_format = fmt
            c.alignment = al or ctr; c.border = bd_thin()

        sm("C", row.get("공고명","-"), al=lft)
        sm("D", bid_no)
        sm("E", row.get("수요/발주기관","-"))
        sm("F", clos)
        sm("G", amt if amt else "", GREEN_L, "#,##0")
        sm("H", fill=YELLOW)
        sm("I", fill=YELLOW)
        sm("J", f"=IFERROR(G{r}/K{r},0)" if amt else "", fmt="#,##0")
        sm("K", fill=YELLOW)
        sm("M", f"=IFERROR(G{r}*N{r+2},0)", fmt="#,##0")   # 기초금액×낙찰하한률
        sm("N", f"=IFERROR(J{r}*N{r+2},0)", fmt="#,##0")
        sm("O", fill=YELLOW)

        # ── P,Q,R: 2행 merge (r+0 ~ r+1) ──────────────────────
        # P,Q,R: merge는 나중에 일괄 처리 (border 먼저)

        # 비고: 하이퍼링크
        c = ws[f"P{r}"]
        c.value = "공고 바로가기" if direct_url else ""
        if direct_url:
            c.hyperlink = direct_url
            c.font = Font(color="0563C1", underline="single")
        c.alignment = ctr

        # 수익: r+2 값을 참조 (merge된 셀에 수식)
        ws[f"Q{r}"].value = f"=IFERROR(M{r}-F{r+2}-G{r+2}-I{r+2}-H{r+2},\"\")"
        ws[f"Q{r}"].number_format = "#,##0"
        ws[f"Q{r}"].font = Font(bold=True); ws[f"Q{r}"].alignment = rgt

        ws[f"R{r}"].value = f"=IFERROR(Q{r}/M{r}*100,\"\")"
        ws[f"R{r}"].number_format = "0.00"
        ws[f"R{r}"].font = Font(bold=True); ws[f"R{r}"].alignment = rgt

        # ── r+1: 서브헤더 (회색) - P,Q,R 이미 merge됨 ──────────
        sub = {
            "C":"1회운반\n예상수량","D":"운행회수","E":"운반비(1회)",
            "F":"운반비합계","G":"총 공인\n계량비","H":"처리비",
            "I":"장비대","M":"부가세","N":"낙찰하한률",# P는 merge phantom — "P":"톤당단가",
        }
        for col in "CDEFGHIJKMN":
            c = ws[f"{col}{r+1}"]
            c.fill = GRAY; c.font = Font(size=8)
            c.border = bd_thin(); c.alignment = ctr
            if col in sub: c.value = sub[col]

        # J, K 서브헤더: calc label 없는 열 → 대각선 border (N/A 표시)
        for col in ["J", "K"]:
            c = ws[f"{col}{r+1}"]
            c.fill = GRAY; c.border = Border(
                left=thin, right=thin, top=thin, bottom=thin,
                diagonal=thin, diagonalDown=True
            )

        ws[f"O{r+1}"].value = "계량방법"; ws[f"O{r+1}"].fill = RED_F
        ws[f"O{r+1}"].font = Font(color="FFFFFF", size=8)
        ws[f"O{r+1}"].alignment = ctr; ws[f"O{r+1}"].border = bd_thin()

        # ── r+2: 계산 행 ─────────────────────────────────────────
        def cc(col, val=None, fill=None, fmt=None):
            c = ws[f"{col}{r+2}"]
            if val  is not None: c.value = val
            if fill is not None: c.fill  = fill
            if fmt  is not None: c.number_format = fmt
            c.alignment = rgt; c.border = bd_thin()
            if not c.fill or c.fill.fgColor.rgb == "00000000": c.fill = WHITE

        cc("C", fill=YELLOW)
        cc("D", f"=IFERROR(K{r}/C{r+2},0)", fmt="#,##0")   # 소수점 제거
        cc("E", fill=YELLOW)
        cc("F", f"=IFERROR(E{r+2}*D{r+2},0)", fmt="#,##0")
        cc("G", f"=IFERROR(D{r+2}*10000,0)", fmt="#,##0")
        cc("H", fill=YELLOW)   # 처리비 직접입력
        cc("I", fill=YELLOW)   # 장비대 직접입력
        # L 제거됨
        ws[f"M{r+2}"].value = "포함"; ws[f"M{r+2}"].alignment = ctr
        ws[f"M{r+2}"].fill = WHITE; ws[f"M{r+2}"].border = bd_thin()

        if lwlt_f is not None:
            cc("N", lwlt_f, fmt="0.00000")
        else:
            cc("N", fill=YELLOW)

        cc("O", fill=YELLOW)   # 계량방법 수동
        # P{r+2}는 merge phantom → 톤당단가 제거
        # Q, R: 3행 merge phantom → 건드리지 않음

        # ── (블록별 medium border는 모든 블록 생성 후 일괄 적용) ──

    # ═══════════════════════════════════════════════════════
    # PASS 1: 전체 데이터 영역에 thin border 일괄 적용
    # ═══════════════════════════════════════════════════════
    from openpyxl.utils import get_column_letter
    n_blocks  = len(sel_df)
    last_row  = SR + n_blocks * BLOCK - 1  # 마지막 데이터 행
    all_merges = {str(m) for m in ws.merged_cells.ranges}

    def is_phantom(row, col):
        """해당 셀이 merge의 phantom(non-top-left)인지 확인"""
        col_letter = get_column_letter(col)
        for m_str in all_merges:
            # B8:B10 형식 파싱
            try:
                from openpyxl.utils import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(m_str)
                if min_row <= row <= max_row and min_col <= col <= max_col:
                    if row != min_row or col != min_col:  # top-left가 아니면 phantom
                        return True
            except Exception:
                pass
        return False

    for ri in range(SR, last_row + 1):
        for ci in COL_RANGE:
            if is_phantom(ri, ci):
                continue
            c = ws.cell(row=ri, column=ci)
            # 대각선 속성 보존
            diag    = c.border.diagonal     if c.border else None
            diag_dn = c.border.diagonalDown if c.border else False
            c.border = Border(
                left=thin, right=thin, top=thin, bottom=thin,
                diagonal=diag, diagonalDown=diag_dn,
            )

    # ═══════════════════════════════════════════════════════
    # PASS 2: No 블록별 외곽에 medium border 덮어씌우기
    # ═══════════════════════════════════════════════════════
    for idx in range(n_blocks):
        r = SR + idx * BLOCK
        r_end = r + BLOCK - 1  # = r+2

        for ri in range(r, r_end + 1):
            for ci in COL_RANGE:
                if is_phantom(ri, ci):
                    continue
                c   = ws.cell(row=ri, column=ci)
                diag    = c.border.diagonal     if c.border else None
                diag_dn = c.border.diagonalDown if c.border else False
                c.border = Border(
                    left   = med if ci == COL_RANGE.start          else c.border.left,
                    right  = med if ci == COL_RANGE.stop - 1       else c.border.right,
                    top    = med if ri == r                         else c.border.top,
                    bottom = med if ri == r_end                     else c.border.bottom,
                    diagonal=diag, diagonalDown=diag_dn,
                )

        # merge top-left 셀들 외곽 medium border 명시 재설정
        ws[f"B{r}"].border = Border(left=med, right=thin, top=med, bottom=med)
        ws[f"P{r}"].border = Border(left=thin, right=thin, top=med, bottom=med)
        ws[f"Q{r}"].border = Border(left=thin, right=thin, top=med, bottom=med)
        ws[f"R{r}"].border = Border(left=thin, right=med,  top=med, bottom=med)

    # ═══════════════════════════════════════════════════════
    # PASS 3: border 설정 완료 후 merge (순서 핵심)
    # border가 regular cell에 있을 때 설정 → merge 후에도 XML에 유지
    # ═══════════════════════════════════════════════════════
    from openpyxl.utils import column_index_from_string
    pc = column_index_from_string("P")
    qc = column_index_from_string("Q")
    rc = column_index_from_string("R")

    for idx in range(len(sel_df)):
        r = SR + idx * BLOCK

        # B열(No): 모든 행에 border 설정
        for ri in range(r, r+3):
            ws.cell(row=ri, column=2).border = Border(
                left   = med,
                right  = thin,
                top    = med  if ri == r   else thin,
                bottom = med  if ri == r+2 else thin,
            )
        # P열(비고): 모든 행에 border 설정
        for ri in range(r, r+3):
            diag    = ws.cell(row=ri, column=pc).border.diagonal
            diag_dn = ws.cell(row=ri, column=pc).border.diagonalDown
            ws.cell(row=ri, column=pc).border = Border(
                left=thin, right=thin,
                top    = med if ri == r   else thin,
                bottom = med if ri == r+2 else thin,
            )
        # Q열(수익): 모든 행에 border 설정
        for ri in range(r, r+3):
            ws.cell(row=ri, column=qc).border = Border(
                left=thin, right=thin,
                top    = med if ri == r   else thin,
                bottom = med if ri == r+2 else thin,
            )
        # R열(수익률): 모든 행에 border 설정
        for ri in range(r, r+3):
            ws.cell(row=ri, column=rc).border = Border(
                left=thin, right=med,
                top    = med if ri == r   else thin,
                bottom = med if ri == r+2 else thin,
            )

        # ★ border 설정 후 merge
        ws.merge_cells(start_row=r, end_row=r+2, start_column=2,  end_column=2)
        ws.merge_cells(start_row=r, end_row=r+2, start_column=pc, end_column=pc)
        ws.merge_cells(start_row=r, end_row=r+2, start_column=qc, end_column=qc)
        ws.merge_cells(start_row=r, end_row=r+2, start_column=rc, end_column=rc)

        # ★ merge 후 top-left + bottom phantom 셀에 border 재설정
        # (openpyxl이 merge 시 phantom top/bottom을 지우므로 재설정 필수)
        # B열 (No)
        ws.cell(row=r,   column=2).border = Border(left=med, right=thin, top=med, bottom=med)
        ws.cell(row=r+2, column=2).border = Border(left=med, right=thin, top=thin, bottom=med)
        # P열 (비고)
        ws.cell(row=r,   column=pc).border = Border(left=thin, right=thin, top=med, bottom=med)
        ws.cell(row=r+2, column=pc).border = Border(left=thin, right=thin, top=thin, bottom=med)
        # Q열 (수익)
        ws.cell(row=r,   column=qc).border = Border(left=thin, right=thin, top=med, bottom=med)
        ws.cell(row=r+2, column=qc).border = Border(left=thin, right=thin, top=thin, bottom=med)
        # R열 (수익률)
        ws.cell(row=r,   column=rc).border = Border(left=thin, right=med, top=med, bottom=med)
        ws.cell(row=r+2, column=rc).border = Border(left=thin, right=med, top=thin, bottom=med)

    # ═══ 낙찰이력 참고 시트 (A방식: 투찰율/낙찰하한율 비율 평균) ═══
    try:
        # 키워드 수집
        all_kws = set()
        for _, row in sel_df.iterrows():
            for k in str(row.get('_키워드','') or '').split(','):
                if k.strip(): all_kws.add(k.strip())

        # 항상 시트 생성 (진단 포함)
        ws_h = wb.create_sheet("낙찰이력참고")
        for col, w in {'A':20,'B':10,'C':16,'D':16,'E':22}.items():
            ws_h.column_dimensions[col].width = w
        ws_h['A1'] = "낙찰이력 참고 (최근 3개월 / A방식)"
        ws_h['A1'].font = Font(bold=True, size=12)
        ws_h.append([])
        ws_h.append(["수집 키워드:", str(sorted(all_kws)) if all_kws else "없음(키워드 미설정)"])
        ws_h.append([])

        if all_kws:
            scsbid = fetch_scsbid_rates_safe(tuple(sorted(all_kws)))
        else:
            scsbid = {}

        if scsbid:
            # 통계 테이블
            ws_h.append(["키워드", "건수", "평균투찰율(%)", "평균낙찰하한율(%)", "참고값(투찰율/낙찰하한율)"])
            for c in ws_h[ws_h.max_row]: c.font = Font(bold=True)
            for kw, info in sorted(scsbid.items(), key=lambda x: -x[1]['count']):
                ws_h.append([
                    kw,
                    info['count'],
                    info.get('avg_t_rate', '-'),
                    info.get('avg_lwlt', '-'),
                    info.get('ratio', '-'),
                ])
            ws_h.append([])
            ws_h.append(["※ 투찰금액 = 기초금액 × 낙찰하한율 × 참고값"])
        else:
            ws_h.append(["⚠️ 낙찰이력 없음", "API 응답 없음 또는 키워드 매칭 공고 없음"])

        # 각 블록 calc행에 참고값 표시
        if scsbid:
            kw_sorted = sorted(all_kws, key=len, reverse=True)
            for idx, (_, row) in enumerate(sel_df.iterrows()):
                r = SR + idx * BLOCK
                bid_kw = next((k for k in kw_sorted
                               if k in str(row.get('공고명',''))), None)
                info = (scsbid.get(bid_kw) if bid_kw else None) or scsbid.get('__전체__')
                if not info: continue
                lwlt = row.get('낙찰하한율', '-')
                try: lwlt_f = float(lwlt) if str(lwlt) not in ('-','','None') else None
                except: lwlt_f = None
                ref = info.get('ratio')
                if ref and lwlt_f and lwlt_f > 0:
                    c = ws.cell(row=r+1, column=15)
                    c.value = '참고값'
                    c.fill = PatternFill("solid", fgColor="FFF2CC")
                    c.font = Font(size=8, bold=True)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = bd_thin()
                    c2 = ws.cell(row=r+2, column=15)
                    c2.value = ref
                    c2.number_format = '0.0000'
                    c2.font = Font(color='FF0000', bold=True)
                    c2.alignment = Alignment(horizontal='right', vertical='center')
                    c2.border = bd_thin()
    except Exception as e:
        try:
            ws_h = wb.create_sheet("낙찰이력참고")
            ws_h['A1'] = f"오류: {str(e)}"
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


@st.cache_data(ttl=600, show_spinner=False)
def fetch_narajangter(keywords, start, end, test_mode):
    service_key = unquote(SERVICE_KEY)
    s_date = start.strftime("%Y%m%d0000")
    e_date = end.strftime("%Y%m%d2359")
    url_base = 'https://apis.data.go.kr/1230000/ad/BidPublicInfoService/'

    all_raw = []
    for kw in keywords:
        try:
            res = requests.get(url_base + 'getBidPblancListInfoServcPPSSrch',
                params={'serviceKey': service_key, 'numOfRows': '50', 'type': 'json',  # 메모리 절약
                        'inqryDiv': '1', 'inqryBgnDt': s_date, 'inqryEndDt': e_date,
                        'bidNtceNm': kw}, timeout=10)
            items = res.json().get('response', {}).get('body', {}).get('items', [])
            if items:
                for it in ([items] if isinstance(items, dict) else items):
                    it['_keyword'] = kw
                    all_raw.append(it)
        except: pass

    if not all_raw: return []
    df_bids = pd.DataFrame(all_raw).drop_duplicates(subset=['bidNtceNo'])

    def call_api(url, params, retries=2):
        """API 호출 + 실패 시 재시도 (속도 초과 대응)"""
        for attempt in range(retries):
            try:
                res = requests.get(url, params=params, timeout=10)
                if res.status_code == 200:
                    return res.json()
            except: pass
            import time; time.sleep(0.3 * (attempt + 1))
        return {}

    def fetch_detail(row):
        b_no  = row['bidNtceNo']
        b_ord = str(row.get('bidNtceOrd', '000')).zfill(3)
        bp = {'ServiceKey': service_key, 'type': 'json', 'inqryDiv': '2',
              'bidNtceNo': b_no, 'bidNtceOrd': b_ord}

        # 지역 조회
        region_val = "⚠️조회실패"
        try:
            r_items = call_api(url_base + 'getBidPblancListInfoPrtcptPsblRgn', bp) \
                .get('response', {}).get('body', {}).get('items', [])
            regs = [str(ri.get('prtcptPsblRgnNm', '')) for ri in
                    ([r_items] if isinstance(r_items, dict) else r_items)
                    if ri.get('prtcptPsblRgnNm')]
            region_val = ", ".join(set(regs)) if regs else "제한없음"
        except: pass

        # 지역 필터: 조회 성공한 경우만 필터 적용, 실패는 포함하되 표시
        if "⚠️조회실패" not in region_val and not region_pass(region_val, test_mode):
            return None

        # 면허 조회
        license_val = "⚠️조회실패"
        try:
            l_items = call_api(url_base + 'getBidPblancListInfoLicenseLimit', bp) \
                .get('response', {}).get('body', {}).get('items', [])
            lics = []
            for li in ([l_items] if isinstance(l_items, dict) else l_items):
                v = li.get('lcnsLmtNm') or li.get('permsnIndstrytyList', '')
                if v: lics.append(str(v))
            license_val = " / ".join(set(lics)) if lics else "제한없음"
        except: pass

        if license_val not in ("⚠️조회실패", "제한없음") and not license_code_pass(license_val, test_mode):
            return None

        # 낙찰방법 화이트리스트 필터 (테스트모드에서는 통과)
        # 통과: 소액수의견적, 적격심사제, 또는 필드 없음/빈값
        ALLOW_BID = ('소액수의견적', '적격심사제')
        sucsfbid_nm = str(row.get('sucsfbidMthdNm', '') or '')
        if not test_mode and sucsfbid_nm and not any(a in sucsfbid_nm for a in ALLOW_BID):
            return None

        return to_row(
            source='나라장터', notice_no=b_no,
            title=row.get('bidNtceNm', '-'), agency=row.get('dminsttNm', '-'),
            notice_dt=row.get('bidNtceDate', row.get('rgstDt', '-')),
            close_dt=row.get('bidClseDt', '-'), open_dt=row.get('opengDt', '-'),
            amount=row.get('asignBdgtAmt', row.get('bdgtAmt', '-')),
            region=region_val, license_info=license_val,
            keyword=row.get('_keyword', '-'),
            # ★ 확인된 URL 패턴으로 직접 조립 (PNPE027_01 고정값 확인됨)
            url=(f"https://www.g2b.go.kr/link/PNPE027_01/single/"
                 f"?bidPbancNo={b_no}"
                 f"&bidPbancOrd={str(row.get('bidNtceOrd','000')).zfill(3)}"),
            # 낙찰하한율: API가 % 단위(예: 89.745)로 줌 → 소수(0.89745)로 변환
            lwlt_rate=round(float(row.get('sucsfbidLwltRate', 0) or 0) / 100, 5)
                      if row.get('sucsfbidLwltRate') else '-',
        )

    results = []
    # 스레드 8개로 제한 (API 초당 30건 제한 대응)
    with ThreadPoolExecutor(max_workers=4) as ex:  # Streamlit Cloud 메모리 절약
        for r in as_completed({ex.submit(fetch_detail, row): row
                                for _, row in df_bids.reset_index(drop=True).iterrows()}):
            try:
                v = r.result()
                if v: results.append(v)
            except: pass
    return results

# =====================================================================
# 3. 국방부 ── 원본(국방부_필터링_완성_최종.py) 로직 100% 반영
#    service_key(언더스코어), 일반입찰 날짜필터 없음
# =====================================================================
@st.cache_data(ttl=600, show_spinner=False)
def fetch_d2b(keywords, start, end, test_mode):
    results   = []
    today_dt  = TODAY
    d2b_start = (today_dt - timedelta(days=10)).strftime("%Y%m%d")
    d2b_end   = (today_dt + timedelta(days=20)).strftime("%Y%m%d")
    target_areas = ["경기도", "평택시", "화성시", "제한없음", "전국"]

    api_configs = [
        {'type': '일반입찰',
         'list_url': 'http://openapi.d2b.go.kr/openapi/service/BidPblancInfoService/getDmstcCmpetBidPblancList',
         'det_url':  'http://openapi.d2b.go.kr/openapi/service/BidPblancInfoService/getDmstcCmpetBidPblancDetail',
         'source': '국방부(일반경쟁)'},
        {'type': '공개수의',
         'list_url': 'http://openapi.d2b.go.kr/openapi/service/BidPblancInfoService/getDmstcOthbcVltrnNtatPlanList',
         'det_url':  'http://openapi.d2b.go.kr/openapi/service/BidPblancInfoService/getDmstcOthbcVltrnNtatPlanDetail',
         'source': '국방부(공개수의)'},
    ]

    for config in api_configs:
        params = {'serviceKey': SERVICE_KEY, 'numOfRows': '500', '_type': 'json'}
        if config['type'] == '공개수의':
            params.update({'prqudoPresentnClosDateBegin': d2b_start,
                           'prqudoPresentnClosDateEnd':   d2b_end})
        try:
            res = requests.get(config['list_url'], params=params, headers=HEADERS, timeout=15)
            if res is None or res.status_code != 200: continue
            items = res.json().get('response', {}).get('body', {}).get('items', {}).get('item', [])
            items = [items] if isinstance(items, dict) else items

            for it in items:
                bid_nm = it.get('bidNm') or it.get('othbcNtatNm', '')
                matched_kw = [kw for kw in keywords if kw in bid_nm]
                if not matched_kw: continue

                p_no    = it.get('pblancNo') or ''
                d_year  = str(it.get('demandYear', ''))
                d_no    = str(it.get('dcsNo', ''))
                p_alpha = "".join([c for c in p_no if c.isalpha()])
                combined_g2b = f"{d_year}{p_alpha}{d_no}"

                p_det = {
                    'service_key': SERVICE_KEY,      # ★ 원본 그대로 언더스코어
                    'pblancNo':    p_no,
                    'pblancOdr':   str(it.get('pblancOdr', '1')).split('.')[0],
                    'demandYear':  d_year,
                    'orntCode':    it.get('orntCode'),
                    'dcsNo':       d_no,
                    '_type':       'json',
                }
                if config['type'] == '공개수의':
                    p_det.update({'ntatPlanDate': it.get('ntatPlanDate'),
                                  'iemNo':        it.get('iemNo')})

                area, budget = "제한없음", 0
                try:
                    det_res  = requests.get(config['det_url'], params=p_det,
                                            headers=HEADERS, timeout=5).json()
                    det_data = det_res.get('response', {}).get('body', {}).get('item', {})
                    if isinstance(det_data, dict):
                        area         = det_data.get('areaLmttList') or "제한없음"
                        combined_g2b = det_data.get('g2bPblancNo') or combined_g2b
                        budget       = (det_data.get('budgetAmount') or
                                        it.get('asignBdgtAmt') or
                                        it.get('budgetAmount') or 0)
                except: pass

                status = it.get('progrsSttus') or "진행중"
                if not ("진행중" in status or status == ""): continue
                if not test_mode and not any(t in area for t in target_areas): continue
                if test_mode or any(t in area for t in target_areas):
                    pass
                else:
                    continue

                results.append(to_row(
                    source=config['source'],
                    notice_no=combined_g2b or p_no or '-',
                    title=bid_nm, agency=it.get('ornt', '-'),
                    notice_dt=format_d2b_dt(it.get('pblancDate', '-')),
                    close_dt=format_d2b_dt(
                        it.get('biddocPresentnClosDt') or it.get('prqudoPresentnClosDt')),
                    open_dt=format_d2b_dt(it.get('opengDt', '-')),
                    amount=int(pd.to_numeric(budget, errors='coerce') or 0),
                    region=area, license_info='미조회',
                    keyword=','.join(matched_kw), url=D2B_HOME,
                ))
        except: pass

    seen, dedup = set(), []
    for r in results:
        key = (r['출처기관'], r['공고번호'])
        if key not in seen:
            seen.add(key); dedup.append(r)
        else:
            for d in dedup:
                if (d['출처기관'], d['공고번호']) == key and r['_키워드'] not in d['_키워드']:
                    d['_키워드'] += f",{r['_키워드']}"
    return dedup

# =====================================================================
# 4. LH
# =====================================================================
@st.cache_data(ttl=600, show_spinner=False)
def fetch_lh(keywords, start, end, test_mode):
    results = []
    try:
        response = requests.get(
            "http://openapi.ebid.lh.or.kr/ebid.com.openapi.service.OpenBidInfoList.dev",
            params={'serviceKey': SERVICE_KEY, 'numOfRows': '500', 'pageNo': '1',
                    'tndrbidRegDtStart': start.strftime("%Y%m%d"),
                    'tndrbidRegDtEnd':   end.strftime("%Y%m%d")}, timeout=20)
        response.encoding = response.apparent_encoding
        clean_xml = re.sub(r'<\?xml.*?>', '', response.text)
        if "<resultCode>00</resultCode>" not in clean_xml: return results
        root = ET.fromstring(f"<root>{clean_xml}</root>")
        for item in root.findall('.//item'):
            title = lh_clean(item.findtext('bidnmKor'))
            matched = [kw for kw in keywords if kw in title]
            if not matched: continue
            rt = [(item.findtext(f'zoneRstrct{n}') or '').strip() for n in range(1, 5)]
            region = " ".join(filter(None, rt))
            if not region_pass(rt, test_mode): continue
            if not lh_license_pass(item, test_mode): continue
            lic = "/".join(sorted(set(
                lh_clean(item.findtext(f'req{n}Reqlic{m}Nm'))
                for n in range(1, 11) for m in range(1, 11)
                if lh_clean(item.findtext(f'req{n}Reqlic{m}Nm'))
            ))) or '제한없음'
            results.append(to_row(
                source='LH', notice_no=item.findtext('bidNum'), title=title,
                agency=lh_clean(item.findtext('zoneHqCd')),
                notice_dt=item.findtext('tndrbidRegDt'),
                close_dt=item.findtext('tndrdocAcptEndDtm'),
                open_dt=item.findtext('openDtm'), amount=item.findtext('fdmtlAmt'),
                region=region if region else '제한없음', license_info=lic,
                keyword=','.join(matched), url=LH_HOME,
            ))
    except: pass
    return results

# =====================================================================
# 5. 가스공사
# =====================================================================
@st.cache_data(ttl=600, show_spinner=False)
def fetch_kogas(keywords, start, end):
    results = []
    try:
        res = requests.get("http://apis.data.go.kr/B551210/bidInfoList/getBidInfoList",
            params={'serviceKey': SERVICE_KEY, 'pageNo': '1', 'numOfRows': '500',
                    'DOCDATE_START': start.strftime("%Y%m%d"),
                    'DOCDATE_END':   end.strftime("%Y%m%d")},
            headers=HEADERS, timeout=15)
        if res.status_code != 200: return results
        sd = start.strftime("%Y%m%d"); ed = end.strftime("%Y%m%d")
        for item in ET.fromstring(res.text).findall('.//item'):
            title = item.findtext('NOTICE_NAME') or '-'
            matched = [kw for kw in keywords if kw in title]
            if not matched: continue
            nd = (item.findtext('NOTICE_DT') or '').replace('-', '')[:8]
            if nd and not (sd <= nd <= ed): continue
            results.append(to_row(
                source='가스공사', notice_no=item.findtext('NOTICE_CODE') or '-',
                title=title, agency=item.findtext('CONT_METHOD_NAME') or '-',
                notice_dt=item.findtext('NOTICE_DT') or '-',
                close_dt=item.findtext('END_DT') or '-', open_dt='-', amount='-',
                region='확인불가', license_info='확인불가',
                keyword=','.join(matched), url=KOGAS_HOME,
            ))
    except: pass
    return results

# =====================================================================
# 6. 수자원공사
# =====================================================================
@st.cache_data(ttl=600, show_spinner=False)
def fetch_kwater(keywords):
    results = []
    search_month = TODAY.strftime('%Y%m')
    for kw in keywords:
        try:
            items = requests.get("http://apis.data.go.kr/B500001/ebid/tndr3/servcList",
                params={'serviceKey': SERVICE_KEY, 'pageNo': '1', 'numOfRows': '100',
                        '_type': 'json', 'searchDt': search_month, 'bidNm': kw},
                headers=HEADERS, timeout=10).json() \
                .get('response', {}).get('body', {}).get('items', {}).get('item', [])
            items = [items] if isinstance(items, dict) else (items or [])
            for it in items:
                title = it.get('tndrPblancNm', '-')
                if kw not in title: continue
                results.append(to_row(
                    source='수자원공사', notice_no=it.get('tndrPbanno', '-'),
                    title=title, agency=it.get('cntrctDeptNm', '-'),
                    notice_dt='-', close_dt=it.get('tndrPblancEnddt', '-'),
                    open_dt='-', amount='-', region='확인불가', license_info='확인불가',
                    keyword=kw,
                    url=(KWATER_DETAIL_BASE + str(it.get('tndrPbanno', '')))
                        if it.get('tndrPbanno') else '-',
                ))
        except: pass
    seen, dedup = set(), []
    for r in results:
        if r['공고번호'] not in seen:
            seen.add(r['공고번호']); dedup.append(r)
    return dedup

# =====================================================================
# 7. 사이드바
# =====================================================================
with st.sidebar:
    st.markdown(f"## 🧹 SWEEP\n**{VERSION}** · 5개 기관 입찰공고 통합 수집")
    st.divider()

    st.markdown("**📅 검색기간** (나라장터·LH·가스공사)")
    date_range = st.date_input("기간", value=(TODAY - timedelta(days=6), TODAY),
                               label_visibility="collapsed")
    start_dt, end_dt = (date_range if isinstance(date_range, tuple) and len(date_range) == 2
                        else (TODAY - timedelta(days=6), TODAY))
    start_dt = datetime.combine(start_dt, datetime.min.time())
    end_dt   = datetime.combine(end_dt,   datetime.min.time())

    st.caption("🏛 국방부: 오늘-10일 ~ 오늘+20일 마감 기준")

    st.markdown("**🎯 키워드**")
    kw_text  = st.text_area("키워드", value=", ".join(DEFAULT_KEYWORDS),
                             height=90, label_visibility="collapsed")
    keywords = tuple(k.strip() for k in kw_text.split(",") if k.strip())

    st.markdown("**🪪 면허코드** (나라장터·국방부)")
    lic_text = st.text_area("면허코드",
                             value=", ".join(DEFAULT_LICENSES),
                             height=55, label_visibility="collapsed",
                             help="쉼표로 구분. 예: 1226, 1227, 6786, 6770")
    OUR_LICENSES = [l.strip() for l in lic_text.split(",") if l.strip()]

    st.markdown("**🪪 LH 면허명** (LH 전용)")
    lh_lic_text = st.text_area("LH면허명",
                                value=", ".join(DEFAULT_LH_LICENSES),
                                height=55, label_visibility="collapsed",
                                help="쉼표로 구분. LH는 면허명(텍스트)으로 매칭")
    WASTE_LICENSE_NAMES = [l.strip() for l in lh_lic_text.split(",") if l.strip()]

    st.markdown("**🏛️ 출처기관**")
    sources  = st.multiselect("기관",
        ["나라장터", "국방부(일반경쟁)", "국방부(공개수의)", "LH", "가스공사", "수자원공사"],
        default=["나라장터", "국방부(일반경쟁)", "국방부(공개수의)", "LH", "가스공사", "수자원공사"],
        label_visibility="collapsed")

    test_mode = st.toggle("🧪 테스트 모드 (필터 OFF)", value=True)

    st.divider()
    st.caption(
        f"{'🔴 필터 OFF — 전체 노출' if test_mode else '🟢 필터 ON'}\n\n"
        f"📍 지역: 경기도·평택·화성·전국\n"
        f"🪪 면허코드: {', '.join(OUR_LICENSES)}\n"
        f"⚠️ 가스공사·수자원공사 지역·면허 미지원"
    )
    st.markdown("")
    run = st.button("🔍 조회 실행", use_container_width=True)

# =====================================================================
# 8. 메인 화면
# =====================================================================
st.title("🧹 SWEEP — 입찰공고 통합 수집")
st.caption(
    f"{VERSION} · 나라장터/국방부/LH/가스공사/수자원공사 · "
    f"공고일 {start_dt:%m/%d}~{end_dt:%m/%d} · "
    f"{'🧪 테스트 모드(필터 OFF)' if test_mode else '✅ 실운영 모드(필터 ON)'}"
)

if "df"     not in st.session_state: st.session_state.df     = pd.DataFrame()
if "errors" not in st.session_state: st.session_state.errors = {}

if run:
    errors, all_rows = {}, []

    steps = {
        "나라장터":   lambda: fetch_narajangter(keywords, start_dt, end_dt, test_mode),
        "국방부":     lambda: fetch_d2b(keywords, start_dt, end_dt, test_mode),
        "LH":        lambda: fetch_lh(keywords, start_dt, end_dt, test_mode),
        "가스공사":   lambda: fetch_kogas(keywords, start_dt, end_dt),
        "수자원공사": lambda: fetch_kwater(keywords),
    }

    prog      = st.progress(0, text="탐지 중...")
    cols      = st.columns(5)
    status_el = {n: c.empty() for n, c in zip(steps, cols)}
    for n in steps:
        status_el[n].markdown(f"**{n}**\n\n⏳")

    # ★ 순차 실행 (메모리 절약 - segfault 방지)
    done = 0
    for name, fn in steps.items():
        try:
            rows = fn()
            err  = None
        except Exception as e:
            rows = []
            err  = str(e)
        done += 1
        prog.progress(done / len(steps), text=f"{name} 완료 ({done}/{len(steps)})")
        if err:
            errors[name] = err
            status_el[name].markdown(f"**{name}**\n\n❌ 오류")
        else:
            all_rows.extend(rows)
            status_el[name].markdown(f"**{name}**\n\n✅ {len(rows)}건")

    prog.progress(1.0, text="완료")
    df = pd.DataFrame(all_rows)
    if not df.empty:
        df = df[df['출처기관'].apply(
            lambda s: any(s == src or s.startswith(src.split('(')[0]) for src in sources))]
    st.session_state.df     = df
    st.session_state.errors = errors

df     = st.session_state.df
errors = st.session_state.get("errors", {})

for name, err in errors.items():
    st.warning(f"⚠️ {name}: {err}")

if df.empty:
    st.info("👈 왼쪽에서 조건을 설정하고 **조회 실행**을 눌러주세요.")
else:
    fail_mask = df['지역제한'].astype(str).str.contains('⚠️조회실패', na=False)
    fail_cnt  = fail_mask.sum()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("총 공고", f"{len(df)}건")
    c2.metric("출처기관", f"{df['출처기관'].nunique()}개")
    c3.metric("지역조회 실패", f"{fail_cnt}건",
              help="API 속도 제한으로 지역 조회에 실패한 공고. 실제 지역 제한이 걸려있을 수 있으니 직접 확인 필요.")
    c4.metric("면허 미확인",
              f"{df['면허정보'].astype(str).str.contains('⚠️조회실패|미조회', na=False).sum()}건")

    if fail_cnt > 0:
        hide_fail = st.checkbox(
            f"⚠️ 지역조회 실패 {fail_cnt}건 숨기기 (직접 확인 필요한 공고)",
            value=False,
            help="나라장터 API 초당 제한으로 지역 조회에 실패했습니다. 충청북도 제천시처럼 "
                 "우리 지역이 아닌 제한이 걸려있을 수 있으니 체크 후 나라장터에서 직접 확인하세요."
        )
        if hide_fail:
            df = df[~fail_mask]
            st.caption(f"조회실패 {fail_cnt}건 제외 → {len(df)}건 표시 중")

    # ── 표시용 컬럼 구성 (내부용 _ 컬럼 숨김, 금액 포맷) ──
    HIDDEN_COLS = ['_공고일시', '_개찰일시', '_키워드', '_비고']
    def fmt_amount(v):
        try:
            n = int(float(str(v).replace(',','').replace('원','').strip()))
            return f"{n:,}원" if n > 0 else '-'
        except:
            return str(v) if v and str(v) not in ('nan','None','-') else '-'

    def fmt_rate(v):
        try:
            return f"{float(v)*100:.3f}%" if v not in ('-', None, '') else '-'
        except:
            return str(v) if v else '-'

    df_editor = df.copy()
    df_editor.insert(0, '선택', False)
    df_editor['금액(원)'] = df_editor['금액(원)'].apply(fmt_amount)
    df_editor['낙찰하한율'] = df_editor['낙찰하한율'].apply(fmt_rate)

    # 숨김 컬럼 제외
    display_cols = [c for c in df_editor.columns if c not in HIDDEN_COLS]
    df_editor = df_editor[display_cols]

    st.divider()
    st.caption("💡 투찰할 공고를 **선택** 후 입찰관리대장으로 다운로드하세요.")

    edited = st.data_editor(
        df_editor,
        use_container_width=True,
        height=540,
        column_config={
            "선택":     st.column_config.CheckboxColumn("선택", width="small"),
            "공고명":   st.column_config.TextColumn("공고명", width="large"),
            "금액(원)": st.column_config.TextColumn("금액(원)"),
            "면허정보": st.column_config.TextColumn("면허정보", width="medium"),
            "상세URL":  st.column_config.LinkColumn("상세URL", display_text="열기"),
        },
        disabled=[c for c in display_cols if c != '선택'],
        hide_index=True,
    )

    selected = edited[edited['선택'] == True]
    n_sel = len(selected)

    col_dl1, col_dl2 = st.columns([1, 1])

    # ── 버튼 1: 전체 결과 엑셀 ──
    with col_dl1:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='통합결과', index=False)
            for src in df['출처기관'].unique():
                df[df['출처기관']==src].to_excel(writer, sheet_name=str(src)[:31], index=False)
        st.download_button("📥 전체 결과 엑셀", data=buf.getvalue(),
            file_name=f"SWEEP_{TODAY:%Y%m%d_%H%M}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

    # ── 버튼 2: 선택 공고 입찰관리대장 ──
    with col_dl2:
        if n_sel == 0:
            st.button(f"📋 입찰관리대장 다운로드 (0건 선택됨)",
                      disabled=True, use_container_width=True)
        else:
            # 선택된 공고번호로 원본 df에서 데이터 가져오기
            sel_nos = selected['공고번호'].tolist()
            sel_df  = df[df['공고번호'].isin(sel_nos)].reset_index(drop=True)
            dajang_buf = make_dajang_excel(sel_df)
            st.download_button(
                f"📋 입찰관리대장 ({n_sel}건 선택)",
                data=dajang_buf,
                file_name=f"입찰관리대장_{TODAY:%Y%m%d_%H%M}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

